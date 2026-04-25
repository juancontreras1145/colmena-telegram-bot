"""Inspecciona libros Excel sin ejecutar macros y sin cargar todo en memoria.

Utiliza openpyxl en modo read_only para perfilar `.xlsx` y `.xlsm`.
Para `.xlsm` se mantiene `keep_vba=False` (default) para nunca tocar las macros.
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import List, Optional
from zipfile import ZipFile, BadZipFile

from openpyxl import load_workbook

from .limits import (
    INFLATED_BYTES_PER_CELL,
    INFLATED_RANGE_RATIO,
    MAX_SHEETS_TO_INSPECT,
)


@dataclass
class SheetProfile:
    name: str
    declared_rows: int
    declared_cols: int
    real_rows: int
    real_cols: int
    non_empty_cells: int
    inflated: bool = False

    @property
    def declared_cells(self) -> int:
        return max(self.declared_rows, 0) * max(self.declared_cols, 0)


@dataclass
class WorkbookProfile:
    path: str
    file_size: int
    is_xlsm: bool
    has_macros: bool
    sheet_count: int
    sheets: List[SheetProfile] = field(default_factory=list)
    inflated_workbook: bool = False
    notes: List[str] = field(default_factory=list)
    best_sheet: Optional[str] = None

    def total_real_cells(self) -> int:
        return sum(s.non_empty_cells for s in self.sheets)


def _detect_macros(path: str, is_xlsm: bool) -> bool:
    """Detecta presencia de macros sin ejecutarlas."""
    if not is_xlsm:
        return False
    try:
        with ZipFile(path) as zf:
            names = zf.namelist()
            return any(n.lower().startswith("xl/vbaproject") for n in names)
    except BadZipFile:
        return False


def _profile_sheet(ws) -> SheetProfile:
    """Recorre una hoja read_only y cuenta filas/cols reales y celdas no vacías."""
    declared_rows = ws.max_row or 0
    declared_cols = ws.max_column or 0

    real_rows = 0
    real_cols = 0
    non_empty = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_has_value = False
        for col_idx, value in enumerate(row, start=1):
            if value is not None and str(value).strip() != "":
                non_empty += 1
                row_has_value = True
                if col_idx > real_cols:
                    real_cols = col_idx
        if row_has_value:
            real_rows = row_idx

    profile = SheetProfile(
        name=ws.title,
        declared_rows=declared_rows,
        declared_cols=declared_cols,
        real_rows=real_rows,
        real_cols=real_cols,
        non_empty_cells=non_empty,
    )
    if profile.declared_cells and profile.non_empty_cells:
        ratio = profile.declared_cells / max(profile.non_empty_cells, 1)
        if ratio >= INFLATED_RANGE_RATIO and profile.declared_cells > 5_000:
            profile.inflated = True
    return profile


def profile_workbook(path: str) -> WorkbookProfile:
    """Perfila un archivo .xlsx o .xlsm en modo read_only y SIN ejecutar macros."""
    file_size = os.path.getsize(path)
    is_xlsm = path.lower().endswith(".xlsm")
    has_macros = _detect_macros(path, is_xlsm)

    profile = WorkbookProfile(
        path=path,
        file_size=file_size,
        is_xlsm=is_xlsm,
        has_macros=has_macros,
        sheet_count=0,
    )

    # keep_vba=False asegura que no se carguen macros aunque el archivo sea .xlsm
    wb = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
        keep_vba=False,
    )
    try:
        profile.sheet_count = len(wb.sheetnames)
        for idx, name in enumerate(wb.sheetnames):
            if idx >= MAX_SHEETS_TO_INSPECT:
                profile.notes.append(
                    f"Se omitieron {profile.sheet_count - MAX_SHEETS_TO_INSPECT} hojas extra "
                    "para no colapsar el servidor."
                )
                break
            ws = wb[name]
            try:
                profile.sheets.append(_profile_sheet(ws))
            except Exception as exc:  # noqa: BLE001
                profile.notes.append(f"No pude leer la hoja '{name}': {exc}")
    finally:
        wb.close()

    # Hoja recomendada: la que tenga más celdas reales.
    if profile.sheets:
        best = max(profile.sheets, key=lambda s: s.non_empty_cells)
        profile.best_sheet = best.name if best.non_empty_cells > 0 else profile.sheets[0].name

    real_cells = profile.total_real_cells()
    if real_cells > 0:
        bytes_per_cell = file_size / real_cells
        if bytes_per_cell > INFLATED_BYTES_PER_CELL and file_size > 1_000_000:
            profile.inflated_workbook = True

    return profile
